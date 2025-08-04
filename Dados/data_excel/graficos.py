from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, BarChart, PieChart

file_path = 'estoque.xlsx'
workbook = load_workbook(file_path)
sheet = workbook['Estoque']

max_row = sheet.max_row

# Gráfico de Barras: Total por Produto

bar_chart = BarChart()
bar_chart.title = "Total de Vendas por Produto"
bar_chart.y_axis.title = "R$"
bar_chart.x_axis.title = "Produto"

# Dados: coluna 'Total'
data = Reference(sheet, min_col=6, min_row=1, max_row=max_row)  # Coluna F (Total)
# Categorias: produtos (coluna A)
categories = Reference(sheet, min_col=1, min_row=2, max_row=max_row)

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
sheet.add_chart(bar_chart, "H2")

# Gráfico de Linha: Lucro por Produto

line_chart = LineChart()
line_chart.title = "Lucro por Produto"
line_chart.y_axis.title = "R$"
line_chart.x_axis.title = "Produto"

data = Reference(sheet, min_col=7, min_row=1, max_row=max_row)  # Coluna G (Lucro)
categories = Reference(sheet, min_col=1, min_row=2, max_row=max_row)

line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(categories)
sheet.add_chart(line_chart, "H20")

# Gráfico de Pizza: Participação no Lucro

pie_chart = PieChart()
pie_chart.title = "Participação no Lucro por Produto"

data = Reference(sheet, min_col=7, min_row=1, max_row=max_row)  # Coluna G (Lucro)
labels = Reference(sheet, min_col=1, min_row=2, max_row=max_row)  # Coluna A (Produto)

pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
sheet.add_chart(pie_chart, "H38")

# Salva o arquivo com os gráficos inseridos
workbook.save(file_path)