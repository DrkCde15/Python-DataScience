from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = 'estoque.xlsx'
workbook = load_workbook(file_path)
sheet = workbook['Estoque']

# Índices das colunas (1-based)
colum_produto = 1
colum_quantidade = 2
colum_preco = 3
colum_preco_venda = colum_preco + 1
colum_total = colum_preco_venda + 1
colum_lucro = colum_total + 1

# Cabeçalhos
sheet.cell(row=1, column=colum_preco_venda, value='Preço de venda')
sheet.cell(row=1, column=colum_total, value='Total')
sheet.cell(row=1, column=colum_lucro, value='Lucro')

max_row = sheet.max_row

for row in range(2, max_row + 1):
    cell_preco = sheet.cell(row=row, column=colum_preco)
    cell_quantidade = sheet.cell(row=row, column=colum_quantidade)
    cell_preco_venda = sheet.cell(row=row, column=colum_preco_venda)
    cell_total = sheet.cell(row=row, column=colum_total)
    cell_lucro = sheet.cell(row=row, column=colum_lucro)

    # Fórmulas
    cell_preco_venda.value = f'={cell_preco.coordinate} * 1.3'
    cell_total.value = f'={cell_preco_venda.coordinate} * {cell_quantidade.coordinate}'
    cell_lucro.value = f'={cell_total.coordinate} - {cell_preco.coordinate} * {cell_quantidade.coordinate}'

# Totais Gerais
linha_total = max_row + 2
sheet.cell(row=linha_total, column=colum_produto, value='Totais Gerais')
sheet.merge_cells(start_row=linha_total, start_column=colum_produto, end_row=linha_total, end_column=colum_quantidade)

sheet.cell(row=linha_total, column=colum_lucro).value = f'=SUM({get_column_letter(colum_lucro)}2:{get_column_letter(colum_lucro)}{max_row})'
sheet.cell(row=linha_total, column=colum_preco).value = f'=SUM({get_column_letter(colum_preco)}2:{get_column_letter(colum_preco)}{max_row})'

workbook.save(file_path)