"""
Criação de planilha Excel com OpenPyXL.

Gera uma planilha de estoque com 50 produtos aleatórios,
incluindo nome, quantidade e preço de cada item.
"""
import random
from openpyxl import Workbook

# ============================================================
# CRIAÇÃO DO WORKBOOK E PLANILHA
# ============================================================

workbook = Workbook()
sheet = workbook.active
sheet.title = "Estoque"

# Cabeçalhos das colunas
headers = ["Produto", "Quantidade", "Preço"]
for colum, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=colum, value=header)

# ============================================================
# GERAÇÃO DE DADOS ALEATÓRIOS
# ============================================================

def gera_produto():
    """Gera um nome de produto aleatório combinando prefixo, tipo e sufixo."""
    prefixos = ['Super', 'Mega', 'Gigante', 'Ultra', 'Power', 'Max']
    tipos = ['Widget', 'Gadget', 'Device', 'Tool', 'Component']
    sufixos = ['Plus', 'Pro', 'X', '2000', 'Elite', 'Prime']
    return f'{random.choice(prefixos)} {random.choice(tipos)} {random.choice(sufixos)}'

num_produtos = 50

for row_num in range(2, num_produtos + 2):
    produto = gera_produto()
    quantidade = random.randint(1, 1000)
    preco = round(random.uniform(10.0, 500.0), 2)
    sheet.cell(row=row_num, column=1, value=produto)
    sheet.cell(row=row_num, column=2, value=quantidade)
    sheet.cell(row=row_num, column=3, value=preco)

# ============================================================
# SALVAMENTO DO ARQUIVO
# ============================================================

file_path = 'estoque.xlsx'
workbook.save(file_path)
print(f"Planilha '{file_path}' criada com {num_produtos} produtos.")